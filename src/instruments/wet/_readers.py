import re
from os.path import getmtime
from pathlib import Path
from typing import Iterable
from typing import List
from typing import Optional
from typing import Set
from typing import Tuple

from src.instruments.wet import CFG_T
from src.instruments.wet import DTA_T
from src.instruments.wet import FP_T
from src.base.log import logger

__all__ = [
    'DTAReader',
    'ConfigReader',
]

log = logger(__name__)


class WETConfigReaderError(Exception):
    pass


class WETConfigValidationError(Exception):
    pass


class Reader:
    _data = None
    _modified: Optional[float] = None
    _fp: Optional[FP_T] = None
    _extensions: Set[str]

    @classmethod
    def read_file(cls, ext: str, fp: FP_T):
        raise NotImplementedError

    @classmethod
    def read(cls, fp: FP_T):
        try:
            _src = Path(fp)
            if not _src.exists():
                raise WETConfigReaderError(f'{fp} does not exist')

            _modified = getmtime(_src)
            if _src != cls._fp or _modified != cls._modified:
                _ext = _src.suffix
                if _ext not in cls._extensions:
                    raise WETConfigReaderError(f'{_ext} bad extension for {cls.__name__}')

                cls._data = cls.read_file(_ext, _src)
                cls._fp, cls._modified = _src, _modified

        except WETConfigReaderError:
            raise

        except Exception as e:
            raise WETConfigReaderError(f'{type(e).__name__}=>{e}')

        else:
            return cls._data


class DTAReader(Reader):
    _data: Optional[DTA_T] = None
    _modified: Optional[float] = None
    _fp: Optional[FP_T] = None
    _extensions = {'.dta', }

    @classmethod
    def read_file(cls, ext: str, fp: FP_T) -> DTA_T:
        _ = ext
        with open(fp, 'rb') as f:
            data = list(map(int, f.read()))
        return [bytes(data[i:i + 271]) for i in range(0, len(data), 271)]


class ConfigReader(Reader):
    _data: Optional[CFG_T] = None
    _modified: Optional[float] = None
    _fp: Optional[FP_T] = None
    _extensions = {'.xlsx', '.csv'}

    RAW_CFG: CFG_T = {(0x5, i): 0x0 for i in range(34, 48)}

    @staticmethod
    def _one_row(d: CFG_T, row: List[str]) -> None:
        target, index, payload = list(map(lambda c: int(c, 16), row))
        if not 10 < index < 256:
            raise WETConfigValidationError(f'{row} failed validation')
        k = target, index
        if k in d:
            raise WETConfigValidationError(f'{k} redeclared in config file')
        d[k] = payload

    @classmethod
    def _parse(cls, num_columns: int, rows: Iterable) -> CFG_T:
        if 3 != num_columns:
            raise WETConfigValidationError('wrong number of columns')
        d = {}
        [cls._one_row(d, row) for row in rows]
        if not d:
            raise WETConfigValidationError('no rows in config')
        return d

    @classmethod
    def read_csv(cls, fp: Path) -> CFG_T:
        f = open(fp, 'r')
        try:
            import csv

            csv_reader = csv.reader(f)
            num_cols = len(next(csv_reader))
            f.seek(0)
            return cls._parse(num_cols, csv_reader)
        finally:
            f.close()

    @classmethod
    def read_xlsx(cls, fp: Path) -> CFG_T:
        from openpyxl import load_workbook

        sheet = load_workbook(
            filename=fp, read_only=True, keep_links=True, keep_vba=False, data_only=True
        ).worksheets[0]
        return cls._parse(sheet.max_row, sheet.iterrows())  # type: ignore

    @classmethod
    def read_file(cls, ext: str, fp: Path) -> Tuple[CFG_T, CFG_T]:
        """
        returns config to write, config to confirm
        """
        # noinspection PyArgumentList
        _data = {
            '.csv': cls.read_csv,
            '.xlsx': cls.read_xlsx,
        }[ext](fp)
        _fp_s = fp.name.lower()
        if 'raw' in _fp_s or 'initial' in _fp_s:
            _w = cls.RAW_CFG.copy()
            _w.update(_data)
            return _w, _data
        else:
            return _data, _data
